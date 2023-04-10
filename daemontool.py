__version__ = '4.1.1'

import os
import logging
import platform
import time
import json
SYSTEM = platform.system()
if SYSTEM == "Windows":
    import winreg
import argparse
from datetime import datetime as dt
import zipfile
from io import BytesIO
from contextlib import contextmanager
import subprocess
import openpyxl
from openpyxl.styles import Border, Side, colors, Font, PatternFill, Alignment
import tqdm
import threading
import functools


BATHEADER = """
1>2# : ^
'''
@echo off
echo Switching stdout to python console!
python "%~f0"
exit /b
rem ^
'''
# This is in python now!

# also you can use this to enter python: @SETLOCAL ENABLEDELAYEDEXPANSION & python -x "%~f0" %* & EXIT /B !ERRORLEVEL!

# @ prevents the script line from being printed
# SETLOCAL ENABLEDELAYEDEXPANSION allows !ERRORLEVEL! to be evaluated after the python script runs
# & allows another command to be run on the same line (similar to UNIX's ;)
# python runs the python interpreter (Must be in %PATH%)
# -x tells python to ignore the first line (Run python -h for details)
# "%~f0" expands to the fully-qualified path of the currently executing batch script (Argument %0). It's quoted in case the path contains spaces
# %* expands all arguments passed to the script, effectively passing them on to the python script
# EXIT /B tells Windows Batch to exit from the current batch file only (Using just EXIT would cause the calling interpreter to exit)
# !ERRORLEVEL! expands to the return code from the previous command after it is run. Used as an argument to EXIT /B, it causes the batch script to exit with the # return code received from the python interpreter

# The batch code is in a multiline string ''' so this is invisible for python.
# The batch parser doesn't see the python code, as it exits before.
# The first line is the key.
# It is valid for batch as also for python!
# In python it's only a senseless compare 1>2 without output, the rest of the line is a comment by the #.
# For batch 1>2# is a redirection of stream 1 to the file 2#.
# The command is a colon : this indicates a label and labeled lines are never printed.
# Then the last caret simply append the next line to the label line, so batch doesn't see the ''' line.

"""

NUITKA_HELP = """
Options:
  --help                show this help message and exit
  --version             Show version information and important details for bug
                        reports, then exit. Defaults to off.
  --module              Create an extension module executable instead of a    
                        program. Defaults to off.
  --standalone          Enable standalone mode for output. This allows you to 
                        transfer the created binary to other machines without 
                        it using an existing Python installation. This also   
                        means it will become big. It implies these option: "--
                        follow-imports" and "--python-flag=no_site". Defaults 
                        to off.
  --onefile             On top of standalone mode, enable onefile mode. This  
                        means not a folder, but a compressed executable is    
                        created and used. Defaults to off.
  --python-debug        Use debug version or not. Default uses what you are   
                        using to run Nuitka, most likely a non-debug version. 
  --python-flag=FLAG    Python flags to use. Default is what you are using to 
                        run Nuitka, this enforces a specific mode. These are  
                        options that also exist to standard Python executable.
                        Currently supported: "-S" (alias "no_site"),
                        "static_hashes" (do not use hash randomization),      
                        "no_warnings" (do not give Python run time warnings), 
                        "-O" (alias "no_asserts"), "no_docstrings" (do not use
                        doc strings), "-u" (alias "unbuffered") and "-m".     
                        Default empty.
  --python-for-scons=PATH
                        If using Python3.3 or Python3.4, provide the path of a
                        Python binary to use for Scons. Otherwise Nuitka can  
                        use what you run Nuitka with or a Python installation 
                        from Windows registry. On Windows Python 3.5 or higher
                        is needed. On non-Windows, Python 2.6 or 2.7 will do  
                        as well.

  Control the inclusion of modules and packages in result:
    --include-package=PACKAGE
                        Include a whole package. Give as a Python namespace,
                        e.g. "some_package.sub_package" and Nuitka will then
                        find it and include it and all the modules found below
                        that disk location in the binary or extension module
                        it creates, and make it available for import by the
                        code. To avoid unwanted sub packages, e.g. tests you
                        can e.g. do this "--nofollow-import-to=*.tests".
                        Default empty.
    --include-module=MODULE
                        Include a single module. Give as a Python namespace,
                        e.g. "some_package.some_module" and Nuitka will then
                        find it and include it in the binary or extension
                        module it creates, and make it available for import by
                        the code. Default empty.
    --include-plugin-directory=MODULE/PACKAGE
                        Include also the code found in that directory,
                        considering as if they are each given as a main file.
                        Overrides all other inclusion options. You ought to
                        prefer other inclusion options, that go by names,
                        rather than filenames, those find things through being
                        in "sys.path". This option is for very special use
                        cases only. Can be given multiple times. Default
                        empty.
    --include-plugin-files=PATTERN
                        Include into files matching the PATTERN. Overrides all
                        other follow options. Can be given multiple times.
                        Default empty.
    --prefer-source-code
                        For already compiled extension modules, where there is
                        both a source file and an extension module, normally
                        the extension module is used, but it should be better
                        to compile the module from available source code for
                        best performance. If not desired, there is --no-
                        prefer-source-code to disable warnings about it.
                        Default off.

  Control the following into imported modules:
    --follow-imports    Descend into all imported modules. Defaults to on in
                        standalone mode, otherwise off.
    --follow-import-to=MODULE/PACKAGE
                        Follow to that module if used, or if a package, to the
                        whole package. Can be given multiple times. Default
                        empty.
    --nofollow-import-to=MODULE/PACKAGE
                        Do not follow to that module name even if used, or if
                        a package name, to the whole package in any case,
                        overrides all other options. Can be given multiple
                        times. Default empty.
    --nofollow-imports  Do not descend into any imported modules at all,
                        overrides all other inclusion options and not usable
                        for standalone mode. Defaults to off.
    --follow-stdlib     Also descend into imported modules from standard
                        library. This will increase the compilation time by a
                        lot and is also not well tested at this time and
                        sometimes won't work. Defaults to off.

  Onefile options:
    --onefile-tempdir-spec=ONEFILE_TEMPDIR_SPEC
                        Use this as a folder to unpack to in onefile mode.
                        Defaults to '%TEMP%/onefile_%PID%_%TIME%', i.e. user
                        temporary directory and being non-static it's removed.
                        Use e.g. a string like
                        '%CACHE_DIR%/%COMPANY%/%PRODUCT%/%VERSION%' which is a
                        good static cache path, this will then not be removed.
    --onefile-child-grace-time=GRACE_TIME_MS
                        When stopping the child, e.g. due to CTRL-C or
                        shutdown, etc. the Python code gets a
                        "KeyboardInterrupt", that it may handle e.g. to flush
                        data. This is the amount of time in ms, before the
                        child it killed in the hard way. Unit is ms, and
                        default 5000.

  Data files:
    --include-package-data=PACKAGE
                        Include data files for the given package name. DLLs
                        and extension modules are not data files and never
                        included like this. Can use patterns the filenames as
                        indicated below. Data files of packages are not
                        included by default, but package configuration can do
                        it. This will only include non-DLL, non-extension
                        modules, i.e. actual data files. After a ":"
                        optionally a filename pattern can be given as well,
                        selecting only matching files. Examples: "--include-
                        package-data=package_name" (all files) "--include-
                        package-data=package_name=*.txt" (only certain type) "
                        --include-package-data=package_name=some_filename.dat
                        (concrete file) Default empty.
    --include-data-files=DESC
                        Include data files by filenames in the distribution.
                        There are many allowed forms. With '--include-data-
                        files=/path/to/file/*.txt=folder_name/some.txt' it
                        will copy a single file and complain if it's multiple.
                        With '--include-data-
                        files=/path/to/files/*.txt=folder_name/' it will put
                        all matching files into that folder. For recursive
                        copy there is a form with 3 values that '--include-
                        data-files=/path/to/scan=folder_name=**/*.txt' that
                        will preserve directory structure. Default empty.
    --include-data-dir=DIRECTORY
                        Include data files from complete directory in the
                        distribution. This is recursive. Check '--include-
                        data-files' with patterns if you want non-recursive
                        inclusion. An example would be '--include-data-
                        dir=/path/some_dir=data/some_dir' for plain copy, of
                        the whole directory. All files are copied, if you want
                        to exclude files you need to remove them beforehand,
                        or use '--noinclude-data-files' option to remove them.
                        Default empty.
    --noinclude-data-files=PATTERN
                        Do not include data files matching the filename
                        pattern given. This is against the target filename,
                        not source paths. So to ignore a file pattern from
                        package data for "package_name" should be matched as
                        "package_name/*.txt". Or for the whole directory
                        simply use "package_name". Default empty.
    --list-package-data=LIST_PACKAGE_DATA
                        Output the data files found for a given package name.
                        Default not done.

  DLL files:
    --noinclude-dlls=PATTERN
                        Do not include DLL files matching the filename pattern
                        given. This is against the target filename, not source
                        paths. So ignore a DLL "someDLL" contained in the
                        package "package_name" it should be matched as
                        "package_name/someDLL.*". Default empty.
    --list-package-dlls=LIST_PACKAGE_DLLS
                        Output the DLLs found for a given package name.
                        Default not done.

  Control the warnings to be given by Nuitka:
    --warn-implicit-exceptions
                        Enable warnings for implicit exceptions detected at
                        compile time.
    --warn-unusual-code
                        Enable warnings for unusual code detected at compile
                        time.
    --assume-yes-for-downloads
                        Allow Nuitka to download external code if necessary,
                        e.g. dependency walker, ccache, and even gcc on
                        Windows. To disable, redirect input from nul device,
                        e.g. "</dev/null" or "<NUL:". Default is to prompt.
    --nowarn-mnemonic=MNEMONIC
                        Disable warning for a given mnemonic. These are given
                        to make sure you are aware of certain topics, and
                        typically point to the Nuitka website. The mnemonic is
                        the part of the URL at the end, without the HTML
                        suffix. Can be given multiple times and accepts shell
                        pattern. Default empty.

  Immediate execution after compilation:
    --run               Execute immediately the created binary (or import the
                        compiled module). Defaults to off.
    --debugger          Execute inside a debugger, e.g. "gdb" or "lldb" to
                        automatically get a stack trace. Defaults to off.
    --execute-with-pythonpath
                        When immediately executing the created binary or
                        module using '--run', don't reset 'PYTHONPATH'
                        environment. When all modules are successfully
                        included, you ought to not need PYTHONPATH anymore,
                        and definitely not for standalone mode.

  Compilation choices:
    --user-package-configuration-file=YAML_FILENAME
                        User provided Yaml file with package configuration.
                        You can include DLLs, remove bloat, add hidden
                        dependencies. Check User Manual for a complete
                        description of the format to use. Can be given
                        multiple times. Defaults to empty.
    --full-compat       Enforce absolute compatibility with CPython. Do not
                        even allow minor deviations from CPython behavior,
                        e.g. not having better tracebacks or exception
                        messages which are not really incompatible, but only
                        different or worse. This is intended for tests only
                        and should *not* be used.
    --file-reference-choice=MODE
                        Select what value "__file__" is going to be. With
                        "runtime" (default for standalone binary mode and
                        module mode), the created binaries and modules, use
                        the location of themselves to deduct the value of
                        "__file__". Included packages pretend to be in
                        directories below that location. This allows you to
                        include data files in deployments. If you merely seek
                        acceleration, it's better for you to use the
                        "original" value, where the source files location will
                        be used. With "frozen" a notation "<frozen
                        module_name>" is used. For compatibility reasons, the
                        "__file__" value will always have ".py" suffix
                        independent of what it really is.
    --module-name-choice=MODE
                        Select what value "__name__" and "__package__" are
                        going to be. With "runtime" (default for module mode),
                        the created module uses the parent package to deduce
                        the value of "__package__", to be fully compatible.
                        The value "original" (default for other modes) allows
                        for more static optimization to happen, but is
                        incompatible for modules that normally can be loaded
                        into any package.

  Output choices:
    --output-filename=FILENAME
                        Specify how the executable should be named. For
                        extension modules there is no choice, also not for
                        standalone mode and using it will be an error. This
                        may include path information that needs to exist
                        though. Defaults to '<program_name>' on this platform.
                        .exe
    --output-dir=DIRECTORY
                        Specify where intermediate and final output files
                        should be put. The DIRECTORY will be populated with
                        build folder, dist folder, binaries, etc. Defaults to
                        current directory.
    --remove-output     Removes the build directory after producing the module
                        or exe file. Defaults to off.
    --no-pyi-file       Do not create a ".pyi" file for extension modules
                        created by Nuitka. This is used to detect implicit
                        imports. Defaults to off.

  Debug features:
    --debug             Executing all self checks possible to find errors in
                        Nuitka, do not use for production. Defaults to off.
    --unstripped        Keep debug info in the resulting object file for
                        better debugger interaction. Defaults to off.
    --profile           Enable vmprof based profiling of time spent. Not
                        working currently. Defaults to off.
    --internal-graph    Create graph of optimization process internals, do not
                        use for whole programs, but only for small test cases.
                        Defaults to off.
    --trace-execution   Traced execution output, output the line of code
                        before executing it. Defaults to off.
    --recompile-c-only  This is not incremental compilation, but for Nuitka
                        development only. Takes existing files and simply
                        compile them as C again. Allows compiling edited C
                        files for quick debugging changes to the generated
                        source, e.g. to see if code is passed by, values
                        output, etc, Defaults to off. Depends on compiling
                        Python source to determine which files it should look
                        at.
    --xml=XML_FILENAME  Write the internal program structure, result of
                        optimization in XML form to given filename.
    --generate-c-only   Generate only C source code, and do not compile it to
                        binary or module. This is for debugging and code
                        coverage analysis that doesn't waste CPU. Defaults to
                        off. Do not think you can use this directly.
    --experimental=FLAG
                        Use features declared as 'experimental'. May have no
                        effect if no experimental features are present in the
                        code. Uses secret tags (check source) per experimented
                        feature.
    --low-memory        Attempt to use less memory, by forking less C
                        compilation jobs and using options that use less
                        memory. For use on embedded machines. Use this in case
                        of out of memory problems. Defaults to off.

  Backend C compiler choice:
    --clang             Enforce the use of clang. On Windows this requires a
                        working Visual Studio version to piggy back on.
                        Defaults to off.
    --mingw64           Enforce the use of MinGW64 on Windows. Defaults to off
                        unless MSYS2 with MinGW Python is used.
    --msvc=MSVC_VERSION
                        Enforce the use of specific MSVC version on Windows.
                        Allowed values are e.g. "14.3" (MSVC 2022) and other
                        MSVC version numbers, specify "list" for a list of
                        installed compilers, or use "latest".  Defaults to
                        latest MSVC being used if installed, otherwise MinGW64
                        is used.
    --jobs=N            Specify the allowed number of parallel C compiler
                        jobs. Defaults to the system CPU count.
    --lto=choice        Use link time optimizations (MSVC, gcc, clang).
                        Allowed values are "yes", "no", and "auto" (when it's
                        known to work). Defaults to "auto".
    --static-libpython=choice
                        Use static link library of Python. Allowed values are
                        "yes", "no", and "auto" (when it's known to work).
                        Defaults to "auto".

  Cache Control:
    --disable-cache=DISABLED_CACHES
                        Disable selected caches, specify "all" for all cached.
                        Currently allowed values are:
                        "all","ccache","bytecode","dll-dependencies". can be
                        given multiple times or with comma separated values.
                        Default none.
    --clean-cache=CLEAN_CACHES
                        Clean the given caches before executing, specify "all"
                        for all cached. Currently allowed values are:
                        "all","ccache","bytecode","dll-dependencies". can be
                        given multiple times or with comma separated values.
                        Default none.
    --disable-bytecode-cache
                        Do not reuse dependency analysis results for modules,
                        esp. from standard library, that are included as
                        bytecode. Same as --disable-cache=bytecode.
    --disable-ccache    Do not attempt to use ccache (gcc, clang, etc.) or
                        clcache (MSVC, clangcl). Same as --disable-
                        cache=ccache.
    --disable-dll-dependency-cache
                        Disable the dependency walker cache. Will result in
                        much longer times to create the distribution folder,
                        but might be used in case the cache is suspect to
                        cause errors. Same as --disable-cache=dll-
                        dependencies.
    --force-dll-dependency-cache-update
                        For an update of the dependency walker cache. Will
                        result in much longer times to create the distribution
                        folder, but might be used in case the cache is suspect
                        to cause errors or known to need an update.

  PGO compilation choices:
    --pgo               Enables C level profile guided optimization (PGO), by
                        executing a dedicated build first for a profiling run,
                        and then using the result to feedback into the C
                        compilation. Note: This is experimental and not
                        working with standalone modes of Nuitka yet. Defaults
                        to off.
    --pgo-args=PGO_ARGS
                        Arguments to be passed in case of profile guided
                        optimization. These are passed to the special built
                        executable during the PGO profiling run. Default
                        empty.
    --pgo-executable=PGO_EXECUTABLE
                        Command to execute when collecting profile
                        information. Use this only, if you need to launch it
                        through a script that prepares it to run. Default use
                        created program.

  Tracing features:
    --report=REPORT_FILENAME
                        Report module, data files, compilation, plugin, etc.
                        details in an XML output file. This is also super
                        useful for issue reporting. Default is off.
    --report-template=REPORT_DESC
                        Report via template. Provide template and output
                        filename "template.rst.j2:output.rst". For built-in
                        templates, check the User Manual for what these are.
                        Can be given multiple times. Default is empty.
    --quiet             Disable all information outputs, but show warnings.
                        Defaults to off.
    --show-scons        Run the C building backend Scons with verbose
                        information, showing the executed commands, detected
                        compilers. Defaults to off.
    --no-progressbar    Disable progress bars. Defaults to off.
    --show-progress     Obsolete: Provide progress information and statistics.
                        Disables normal progress bar. Defaults to off.
    --show-memory       Provide memory information and statistics. Defaults to
                        off.
    --show-modules      Provide information for included modules and DLLs
                        Obsolete: You should use '--report' file instead.
                        Defaults to off.
    --show-modules-output=PATH
                        Where to output '--show-modules', should be a
                        filename. Default is standard output.
    --verbose           Output details of actions taken, esp. in
                        optimizations. Can become a lot. Defaults to off.
    --verbose-output=PATH
                        Where to output from '--verbose', should be a
                        filename. Default is standard output.

  General OS controls:
    --disable-console   When compiling for Windows or macOS, disable the
                        console window and create a GUI application. Defaults
                        to off.
    --enable-console    When compiling for Windows or macOS, enable the
                        console window and create a console application. This
                        disables hints from certain modules, e.g. "PySide"
                        that suggest to disable it. Defaults to true.
    --force-stdout-spec=FORCE_STDOUT_SPEC
                        Force standard output of the program to go to this
                        location. Useful for programs with disabled console
                        and programs using the Windows Services Plugin of
                        Nuitka commercial. Defaults to not active, use e.g.
                        '%PROGRAM%.out.txt', i.e. file near your program.
    --force-stderr-spec=FORCE_STDERR_SPEC
                        Force standard error of the program to go to this
                        location. Useful for programs with disabled console
                        and programs using the Windows Services Plugin of
                        Nuitka commercial. Defaults to not active, use e.g.
                        '%PROGRAM%.err.txt', i.e. file near your program.

  Windows specific controls:
    --windows-icon-from-ico=ICON_PATH
                        Add executable icon. Can be given multiple times for
                        different resolutions or files with multiple icons
                        inside. In the later case, you may also suffix with
                        #<n> where n is an integer index starting from 1,
                        specifying a specific icon to be included, and all
                        others to be ignored.
    --windows-icon-from-exe=ICON_EXE_PATH
                        Copy executable icons from this existing executable
                        (Windows only).
    --onefile-windows-splash-screen-image=SPLASH_SCREEN_IMAGE
                        When compiling for Windows and onefile, show this
                        while loading the application. Defaults to off.
    --windows-uac-admin
                        Request Windows User Control, to grant admin rights on
                        execution. (Windows only). Defaults to off.
    --windows-uac-uiaccess
                        Request Windows User Control, to enforce running from
                        a few folders only, remote desktop access. (Windows
                        only). Defaults to off.

  macOS specific controls:
    --macos-target-arch=MACOS_TARGET_ARCH
                        What architectures is this to supposed to run on.
                        Default and limit is what the running Python allows
                        for. Default is "native" which is the architecture the
                        Python is run with.
    --macos-create-app-bundle
                        When compiling for macOS, create a bundle rather than
                        a plain binary application. Currently experimental and
                        incomplete. Currently this is the only way to unlock
                        disabling of console.Defaults to off.
    --macos-app-icon=ICON_PATH
                        Add icon for the application bundle to use. Can be
                        given only one time. Defaults to Python icon if
                        available.
    --macos-signed-app-name=MACOS_SIGNED_APP_NAME
                        Name of the application to use for macOS signing.
                        Follow "com.YourCompany.AppName" naming results for
                        best results, as these have to be globally unique, and
                        will potentially grant protected API accesses.
    --macos-app-name=MACOS_APP_NAME
                        Name of the product to use in macOS bundle
                        information. Defaults to base filename of the binary.
    --macos-app-mode=MODE
                        Mode of application for the application bundle. When
                        launching a Window, and appearing in Docker is
                        desired, default value "gui" is a good fit. Without a
                        Window ever, the application is a "background"
                        application. For UI elements that get to display
                        later, "ui-element" is in-between. The application
                        will not appear in dock, but get full access to
                        desktop when it does open a Window later.
    --macos-sign-identity=MACOS_APP_VERSION
                        When signing on macOS, by default an ad-hoc identify
                        will be used, but with this option your get to specify
                        another identity to use. The signing of code is now
                        mandatory on macOS and cannot be disabled. Default
                        "ad-hoc" if not given.
    --macos-sign-notarization
                        When signing for notarization, using a proper TeamID
                        identity from Apple, use the required runtime signing
                        option, such that it can be accepted.
    --macos-app-version=MACOS_APP_VERSION
                        Product version to use in macOS bundle information.
                        Defaults to "1.0" if not given.
    --macos-app-protected-resource=RESOURCE_DESC
                        Request an entitlement for access to a macOS protected
                        resources, e.g.
                        "NSMicrophoneUsageDescription:Microphone access for
                        recording audio." requests access to the microphone
                        and provides an informative text for the user, why
                        that is needed. Before the colon, is an OS identifier
                        for an access right, then the informative text. Legal
                        values can be found on https://developer.apple.com/doc
                        umentation/bundleresources/information_property_list/p
                        rotected_resources and the option can be specified
                        multiple times. Default empty.

  Linux specific controls:
    --linux-icon=ICON_PATH
                        Add executable icon for onefile binary to use. Can be
                        given only one time. Defaults to Python icon if
                        available.

  Binary Version Information:
    --company-name=COMPANY_NAME
                        Name of the company to use in version information.
                        Defaults to unused.
    --product-name=PRODUCT_NAME
                        Name of the product to use in version information.
                        Defaults to base filename of the binary.
    --file-version=FILE_VERSION
                        File version to use in version information. Must be a
                        sequence of up to 4 numbers, e.g. 1.0 or 1.0.0.0, no
                        more digits are allowed, no strings are allowed.
                        Defaults to unused.
    --product-version=PRODUCT_VERSION
                        Product version to use in version information. Same
                        rules as for file version. Defaults to unused.
    --file-description=FILE_DESCRIPTION
                        Description of the file used in version information.
                        Windows only at this time. Defaults to binary
                        filename.
    --copyright=COPYRIGHT_TEXT
                        Copyright used in version information. Windows only at
                        this time. Defaults to not present.
    --trademarks=TRADEMARK_TEXT
                        Copyright used in version information. Windows only at
                        this time. Defaults to not present.

  Plugin control:
    --enable-plugin=PLUGIN_NAME
                        Enabled plugins. Must be plug-in names. Use '--plugin-
                        list' to query the full list and exit. Default empty.
    --disable-plugin=PLUGIN_NAME
                        Disabled plugins. Must be plug-in names. Use '--
                        plugin-list' to query the full list and exit. Most
                        standard plugins are not a good idea to disable.
                        Default empty.
    --plugin-no-detection
                        Plugins can detect if they might be used, and the you
                        can disable the warning via "--disable-plugin=plugin-
                        that-warned", or you can use this option to disable
                        the mechanism entirely, which also speeds up
                        compilation slightly of course as this detection code
                        is run in vain once you are certain of which plugins
                        to use. Defaults to off.
    --plugin-list       Show list of all available plugins and exit. Defaults
                        to off.
    --user-plugin=PATH  The file name of user plugin. Can be given multiple
                        times. Default empty.
    --show-source-changes
                        Show source changes to original Python file content
                        before compilation. Mostly intended for developing
                        plugins. Default False.
"""


def desktop_path():
    """return your desktop path"""
    if SYSTEM == "Windows":
        key = winreg.OpenKey(winreg.HKEY_CURRENT_USER, r'Software\Microsoft\Windows\CurrentVersion\Explorer\Shell Folders')
        return winreg.QueryValueEx(key, "Desktop")[0]
    else:
        return "/home/"


def win_command(command):
    """get windows command response"""
    sub = subprocess.Popen(f"{command}", stdout=subprocess.PIPE, stderr=subprocess.PIPE, encoding='utf-8')
    out, err = sub.communicate()
    return out, err


def bash_command(command):
    """get linux bash command response"""
    response = []
    p = subprocess.Popen(command, shell=True, stdout=subprocess.PIPE, stderr=subprocess.STDOUT)
    for line in iter(p.stdout.readline, b''):
        line = line.decode().strip()
        if line:
            response.append(line)
            print(line)
    return response


def get_path(path):
    """generate all abs path under the given path"""
    for dirname, _, filenames in os.walk(path):
        for filename in filenames:
            yield os.path.join(dirname, filename)


def get_all_path(rootdir):
    """return all files abs paths in the given directory"""
    path_list = []
    all_list = os.listdir(rootdir)
    for i in range(len(all_list)):
        com_path = os.path.join(rootdir, all_list[i])
        if os.path.isfile(com_path):
            path_list.append(com_path)
        if os.path.isdir(com_path):
            path_list.extend(get_all_path(com_path))
    return path_list


def folder_level_X_path(path, level=3):
    """return the level of the path under the given path"""
    path_list = []
    all_list = os.listdir(path)
    for i in range(len(all_list)):
        try:
            com_path = os.path.join(path, all_list[i])
        except NotADirectoryError:
            continue
        if level == 1:
            path_list.append(com_path)
        else:
            path_list.extend(folder_level_X_path(com_path, level=level-1))
    return path_list


def get_runtime_path():
    """return Current directory path"""
    return os.getcwd()


def join_path(path, file):
    """fake os.path.join method"""
    return os.path.join(path, file)


def get_current_time(format="%Y-%m-%d %H:%M:%S"):
    """
    return Current time format in string
    default format: %Y-%m-%d %H:%M:%S
    """
    return time.strftime(format, time.localtime())


def __logorder__(func):
    """A wrapper for mylogging"""
    def wrapper(self, msg):
        if self.showlog:
            if not self.savelog:
                if self.branch:
                    getattr(logging, func.__name__)(msg=f"[{self.branch}] - {msg}")
                else:
                    getattr(logging, func.__name__)(msg=f" {msg}")
            else:
                if self.branch:
                    getattr(self.logger, func.__name__)(msg=f"[{self.branch}] - {msg}")
                else:
                    getattr(self.logger, func.__name__)(msg=f" {msg}")
        else:
            pass
        return func(self, msg)
    return wrapper


class mylogging():
    """
        A simpl logging system
        >>> usage:  
            log = mylogging()
            log.info
            log.warning
            log.error
            log.debug 
            ...
        
        branch : is the log branch name in logging
        llevel : is the shown log level in logging
        showlog: the switch to enable log or not
    """

    level_relation = {
        'debug'  : logging.DEBUG,
        'info'   : logging.INFO,
        'warning': logging.WARNING,
        'error'  : logging.ERROR
    }

    format = '%(asctime)s [%(levelname)s]%(message)s'

    def __init__(self, branch=None, llevel='debug', showlog=True, savelog=None):
        self.showlog = showlog
        self.branch = branch
        self.savelog = savelog

        if not savelog:
            logging.basicConfig(level=self.level_relation[llevel], format=self.format)
        else:
            os.remove(savelog)
            fh = logging.FileHandler(savelog)
            sh = logging.StreamHandler()
            ft = logging.Formatter(self.format)
            fh.setLevel(self.level_relation[llevel])
            sh.setLevel(self.level_relation[llevel])
            fh.setFormatter(ft)
            sh.setFormatter(ft)

            self.logger = logging.getLogger()
            self.logger.setLevel(self.level_relation[llevel])
            self.logger.addHandler(fh)
            self.logger.addHandler(sh)

    @__logorder__
    def info(self, msg):
        pass
    
    @__logorder__
    def debug(self, msg):
        pass
    
    @__logorder__
    def warning(self, msg):
        pass
    
    @__logorder__
    def error(self, msg):
        pass

    @__logorder__
    def exception(self, msg):
        pass


def timethis(func):
    """A wrapper for counting functions time spent"""
    def wrapper(*args, **kwargs):
        start = time.time()
        result = func(*args, **kwargs)
        end = round((time.time() - start), 3)
        print(f'{func.__name__} running time: {end}sec.')
        return result
    return wrapper


class CodeTimer(object):
    """
        Class for counting functions time spent\n
        >>> example:
            with CodeTimer():
                code line;
                code line;
                ...
    """
    def __init__(self, keep_num=3):
        self.start = time.time()
        self.keep_num = keep_num

    def __enter__(self):
        return self

    def __exit__(self, *_):
        self.stop = time.time()
        self.cost = self.stop - self.start
        print(f'Running time: {round(self.cost, self.keep_num)}sec')


def mkdir(path):
    """make directorys for the given path"""
    path = path.strip().rstrip("\\")
    if not os.path.exists(path):
        os.makedirs(path)
    else:
        print(f'{path} existed!')


def dict2json(target_dict, json_name, json_path):
    """"dict to json"""
    file = join_path(json_path, f'{json_name}.json')
    if not os.path.exists(json_path):
        mkdir(json_path)
    content = json.dumps(target_dict, indent=4)
    with open(file, 'w') as json_file:
        json_file.write(content)


def json2dict(json_path):
    """"json to dict"""
    with open(json_path, 'r', encoding='UTF-8') as f:
        return json.load(f)


def parserinit(description, *args:dict):
    """
        add arguments for scripts
        >>> use:
            args = parserinit(
            'Script description there',
            {'param': '-arg1', 'help': 'help1'},
            {'param': '-arg2', 'help': 'help2'},
            ....)
        then you can get arg by args.args1, args.args2 ...
    """
    parser = argparse.ArgumentParser(description=description)
    for arg in args:
        if 'param' and 'help' in arg.keys():
            parser.add_argument(arg['param'], help=arg['help'])
        else:
            raise KeyError("Wrong arguments, arg['param'] and arg['help] must need")
    return parser.parse_args()


class zipreader(object):
    """
        zipreader
        open a zip and return a file content

        args:
            zippath is the path of the zip file
            filekeyword is the file path in the zip you want to open

        >>> use:
            with zipreader(zippath, filekeyword) as z:
        z is content list now
    """
    def __init__(self, zippath, filekeyword):
        with zipfile.ZipFile(zippath, "r") as z:
            for zipfile_path in z.namelist():
                if filekeyword in zipfile_path:
                    with z.open(zipfile_path, 'r') as file:
                        self.content = list(map(lambda x: x.decode(), file.readlines()))

    def __enter__(self):
        return self.content

    def __exit__(self, *_):
        pass


class zip2reader(object):
    """
        open a zip which is inside zip and return a file content

        args:
            zippath is the path of the zip file
            subziptype is the type of the zip inside the zip file
            filekeyword is the file path in the zip you want to open

        >>> use:
            with zip2reader(zippath, subziptype='zip', filekeyword='.log') as z:
        z is content list now
    """

    def __init__(self, zippath, subziptype='zip', filekeyword=''):
        with zipfile.ZipFile(zippath, "r") as mainzip:
            for mainzipcontent in [i for i in mainzip.namelist() if f".{subziptype}" in i]:
                subzip = BytesIO(mainzip.read(mainzipcontent))
                with zipfile.ZipFile(subzip, "r") as sz:
                    for subzipcontent in [i for i in sz.namelist() if filekeyword in i]:
                            with sz.open(subzipcontent, 'r') as target:
                                self.content = list(map(lambda x: x.decode(), target.readlines()))
    def __enter__(self):
        return self.content

    def __exit__(self, *_):
        pass


class DateTransformer():
    """
        Date string transformation
    """
    def __init__(self, datestring):
        datestring = datestring.replace('-','')
        _FormatDateString = dt.strptime(datestring,"%Y%m%d")
        _DateInformation = _FormatDateString.isocalendar()
        
        self.year        = int(_DateInformation[0])
        self.week        = int(_DateInformation[1])
        self.month       = int(_FormatDateString.month)
        self.quarter     = int(self.month // 4 + 1)
        if len(str(self.week)) == 2:
            self.yearweek  = f"{self.year}W{self.week}"
        elif len(str(self.week)) == 1:
            self.yearweek  = f"{self.year}W0{self.week}"
        if len(str(self.month)) == 2:
            self.yearmonth = f"{self.year}M{self.month}"
        elif len(str(self.month)) == 1:
            self.yearmonth = f"{self.year}M0{self.month}"
        self.yearquarter = f"{self.year}Q{self.quarter}"
        self.timestamp     = int(time.mktime(time.strptime(datestring,"%Y%m%d")))


@contextmanager
def ignored(exception=Exception, func=lambda:None, **kwargs):
    """
        Use: 
        with ignored(func = func, **kwargs):
            xx 
        same as 
        try: 
            xx 
        except Exception
            func
    """
    try:
        yield
    except exception:
        func(**kwargs)


class xlsxDesigner():
    """
        Generate an openpyxl type xlsx design
    """

    sggcolor = {
        'BlueAngel'  : 'B7CEEC',
        'MagicMint'  : 'AAF0D1',
        'CreamWhite' : 'FFFDD0',
        'PeachPink'  : 'F98B88',
        'PeriWinkle' : 'CCCCFF'
    }

    def __init__(self, bgcolor="BlueAngel", hzalign="left", font='Candara Light', fontsize='14', fontbold=False):

        self.border = Border(
            top    = Side(border_style='thin', color=colors.BLACK),
            bottom = Side(border_style='thin', color=colors.BLACK),
            left   = Side(border_style='thin', color=colors.BLACK),
            right  = Side(border_style='thin', color=colors.BLACK)
        )
        self.font = Font(font, size=fontsize, bold=fontbold)
        try:
            self.fill = PatternFill('solid', fgColor=self.sggcolor[bgcolor]) 
        except KeyError:
            self.fill = PatternFill('solid', fgColor=bgcolor)
        self.alignment = Alignment(horizontal=hzalign, vertical='center') # left, general, right, center


class xlsxMaker():
    """
        A class for making a xlsx file with openpyxl extension
    """

    def __init__(self):
        self.wb = openpyxl.Workbook()
        self.wb.remove(self.wb['Sheet'])

    def create_sheet(self, sheetname='undefine'):
        return self.wb.create_sheet(sheetname)

    def _get_num_colnum_dict(self):
        num_str_dict = {}
        A_Z = [chr(a) for a in range(ord('A'), ord('Z') + 1)]
        AA_AZ = ['A' + chr(a) for a in range(ord('A'), ord('Z') + 1)]
        A_AZ = A_Z + AA_AZ
        for i in A_AZ:
            num_str_dict[A_AZ.index(i) + 1] = i
        return num_str_dict

    def auto_fit_width(self, excel_name:str, sheet_name:str):
        wb = openpyxl.load_workbook(excel_name)
        sheet = wb[sheet_name]
        max_column = sheet.max_column
        max_row = sheet.max_row
        max_column_dict = {}
        num_str_dict = self._get_num_colnum_dict()
        for i in range(1, max_column + 1):
            for j in range(1, max_row + 1):
                column = 0
                sheet_value = sheet.cell(row=j, column=i).value
                sheet_value_list = [k for k in str(sheet_value)]
                for v in sheet_value_list:
                    if v.isdigit() == True or v.isalpha() == True:
                        column += 1.1
                    else:
                        column += 2.2
                try:
                    if column > max_column_dict[i]:
                        max_column_dict[i] = column
                except Exception as e:
                    max_column_dict[i] = column
        for key, value in max_column_dict.items():
            sheet.column_dimensions[num_str_dict[key]].width = value + 2
        wb.save(excel_name)

    def wirte2cell(self, sheet, design, row, column, value, fill=False):
        sheet.cell(row=row, column=column).value       = value
        sheet.cell(row=row, column=column).border      = design.border
        sheet.cell(row=row, column=column).font        = design.font
        sheet.cell(row=row, column=column).alignment   = design.alignment
        if fill:
            sheet.cell(row=row, column=column).fill   = design.fill
    
    def write2mergecell(self, sheet, design, start_row, end_row, start_column, end_column, value, fill=False):
        self.wirte2cell(sheet, design, start_row, start_column, value, fill)
        sheet.merge_cells(start_row=start_row, start_column=start_column, end_row=end_row, end_column=end_column)
    
    def save(self, xlsxname, xlsxpath, allautowidth=True):
        self.wb.save(f"{xlsxpath}{SEP}{xlsxname}.xlsx")
        if allautowidth:
            for sheet in self.wb.sheetnames:
                self.auto_fit_width(excel_name=f"{xlsxpath}{SEP}{xlsxname}.xlsx", sheet_name=sheet)


class NuitkaMake():
    """
        usage:
            use Nuitka to build app

        >>> example:
            nm = NuitkaMake("main.py")
            nm.ADD_ARG('onefile')
            nm.ADD_ARG('standalone')
            nm.ADD_ARG('remove-output')
            nm.ADD_ARG('follow-imports')
            nm.ADD_ARG(f'output-filename="{EXE}"')
            nm.ADD_ARG(f'output-dir="{CWD}"')
            nm.ADD_ARG(f'windows-icon-from-ico="{ICON}"')
            nm.ADD_ARG('file-description="Test Time Counter"')
            nm.ADD_ARG('copyright="Copyright (C) 2023 NVIDIA. all right reserved."')
            nm.ADD_ARG(f'file-version="{VER}"')
            nm.ADD_ARG(f'product-version="{VER}"')
            nm.MAKE()
    """
    def __init__(self, main):
        if SYSTEM == 'Windows':
            self.command = 'python -m nuitka'
        elif SYSTEM == 'Linux':
            self.command = 'python3 -m nuitka'
        else:
            raise TypeError('Unsupported system')
        self.main = main

    def ADD_ARG(self, arg):
        self.command = self.command.replace(f' {self.main}', '')
        self.command = f'{self.command} --{arg} {self.main}'
        print(f'Adding arg: --{arg}')

    def MAKE(self):
        print("Nuitka building start ...")
        with CodeTimer():
          os.system(self.command)
    
    def HELP(self):
        print(NUITKA_HELP)


class Py2BAT():
    """
        usage:
            make py file to windows batch file

        args:
            batchname  : name of your batch file
            output_path: define the generated batch file path
        
        example:
            Py2BAT("main.py", batname='test').MAKE()

    """
    def __init__(self, main, batname="Null", output_path=get_runtime_path()):
        self.main = main
        self.batname = batname
        self.output_path = output_path
        
    def MAKE(self):
        with CodeTimer():
            with open(f'{self.main}','r', encoding="utf8") as script:
                codes = script.readlines()
                with open(f"{self.output_path}{SEP}{self.batname}.bat", 'w', encoding="utf8") as batch:
                    batch.write(BATHEADER)
                    batch.writelines(codes)


def _progress_bar(function, estimated_time, tstep, progress_name, tqdm_kwargs={}, args=[], kwargs={}):
    """Tqdm wrapper for a long-running function
        >>> args:
            function       - function to run
            estimated_time - how long you expect the function to take
            tstep          - time delta (seconds) for progress bar updates
            tqdm_kwargs    - kwargs to construct the progress bar
            args           - args to pass to the function
            kwargs         - keyword args to pass to the function
        
        ret:
            function(*args, **kwargs)
        
        >>> example:
                test = _progress_bar(
                    running_function,
                    estimated_time=5, 
                    tstep=1/5.0,
                    tqdm_kwargs={"bar_format":"{desc}{percentage:3.1f}%|{bar:25}|"},
                    args=(1, "foo"), 
                    kwargs={"spam":"eggs"}
                    )
    """
    ret = [None]  # Mutable var so the function can store its return value
    pbar = tqdm.tqdm(total=estimated_time,**tqdm_kwargs)
    pbar.set_description(progress_name)


    def return_save(function, ret, *args, **kwargs):
        ret[0] = function(*args, **kwargs)
    
    class _progress_bar_build_in_print():
        @staticmethod
        def print_with_bar(msg):
            pbar.bar_format = tqdm_kwargs["bar_format"] + msg
        @staticmethod
        def print_in_line(msg):
            pbar.write(msg)
        print = print_with_bar
        write = print_in_line

    if '_progress_bar' in kwargs.keys():
        kwargs['_progress_bar'] = _progress_bar_build_in_print

    thread = threading.Thread(target=return_save, args=(function, ret) + tuple(args), kwargs=kwargs)
    actuall_time = 0
    thread.start()

    while thread.is_alive():
        thread.join(timeout=tstep)
        if actuall_time < estimated_time:
            pbar.update(tstep)
            actuall_time += tstep
        # for actual running time are longer than estimated_time
        if actuall_time + tstep > estimated_time:
            tstep = estimated_time - actuall_time
            pbar.update(tstep)
            actuall_time += tstep
    
    # for actual function running time is shorter than estimated_time
    while(actuall_time < estimated_time):
        if actuall_time + tstep > estimated_time:
            tstep = estimated_time - actuall_time
        pbar.update(tstep)
        actuall_time += tstep

    pbar.close()
    return ret[0]


def progressbar(estimated_time, tstep=0.1, progress_name='',tqdm_kwargs={"bar_format":"{desc}{percentage:3.0f}%|{bar:25}|"}):
    """
        Decorate a function to add a progress bar

        >>> example:
            @progress_wrapped(estimated_time=8, tstep=0.2, progress_name='test')
            def arunning_function(*args, **kwargs):
                pass
            
            there provide a build in bar-print-function if your fucntion have a "_progress_bar" parameter

            your can assign _progress_bar to anthing

            it will be redirected to build in bar-print-function

            then you can use 
            _progress_bar.print_with_bar(message)(or _progress_bar.print): 
                    print message with bar

            _progress_bar.print_in_line(message)(or _progress_bat.write) : 
                    print message in another line but keep progress bar moving
        
        >>> example:
                class A():
                    @staticmethod
                    @progressbar(estimated_time=8, tstep=0.1, progress_name='this is a test')
                    def test_print(_progress_bar):
                        import time
                        _progress_bar.print("test1")
                        _progress_bar.write("test1")
                        time.sleep(2)
                        _progress_bar.print_with_bar("test2")
                        _progress_bar.print_in_line("test2")
                        time.sleep(2)
                        _progress_bar.print_with_bar("test3")
                        _progress_bar.print_in_line("test3")
                test_bar = None
                A.test_print(_progress_bar=test_bar)
    """
    # back up: tqdm_kwargs={"bar_format":"{desc}: {percentage:3.0f}%|{bar:25}| {n:.1f}/{total:.1f} [{elapsed}<{remaining}]"}
    def real_decorator(function):
        @functools.wraps(function)
        def wrapper(*args, **kwargs):
            return _progress_bar(function, estimated_time=estimated_time, tstep=tstep, progress_name=progress_name, tqdm_kwargs=tqdm_kwargs, args=args, kwargs=kwargs)
        return wrapper
    return real_decorator


SEP            =  os.sep
DESKTOPPATH    =  desktop_path()
CURRENTTIME    =  get_current_time()
CURRENTWORKDIR =  get_runtime_path()
CURRENTYEAR    =  int(dt.now().isocalendar()[0])
CURRENTWEEK    =  int(dt.now().isocalendar()[1])
SYSTEM         =  platform.system()


def GV_list():
    print("SEP            : system file path divided sign")
    print("DESKTOPPATH    : your desktop path")
    print("CURRENTTIME    : current date")
    print("CURRENTWORKDIR : current work directory")
    print("CURRENTYEAR    : current year")
    print("CURRENTWEEK    : current week")


if __name__ == '__main__':
    daemontool_log = mylogging(branch='DAEMON_LOGGING')
    daemontool_log.info(f'daemontool - v{__version__}')
